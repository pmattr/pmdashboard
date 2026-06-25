from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()

HEADER_BG = "1e3a5f"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "e8f0fe"
ARIAL = "Arial"

def style_header(cell):
    cell.font = Font(name=ARIAL, bold=True, color=HEADER_FG)
    cell.fill = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_cell(cell, row_idx, wrap=False):
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", start_color=ALT_ROW_BG, end_color=ALT_ROW_BG)
    cell.font = Font(name=ARIAL, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)

def auto_size_columns(ws, min_width=8, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                cell_max = max(len(line) for line in lines)
                if cell_max > max_len:
                    max_len = cell_max
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
blue_fill  = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ── Sheet 1: Rows ──────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Rows"
ws1.freeze_panes = "A2"

col_names1 = ["ID", "Environment", "Title", "Status", "Target Date", "Owner", "ADO Ref", "Notes"]
for col_idx, name in enumerate(col_names1, 1):
    style_header(ws1.cell(row=1, column=col_idx, value=name))

rows_data = [
    (1,"QA","Threshold Merge for Basis Content","Done","Jan 27 2026","NPD/NER","User Story 1448457",""),
    (2,"QA","Load full content set using v4.7 - qa","Done","Feb 26 2026","Dev","",""),
    (3,"QA","Run all non-basis content (200 records) against V4.8","Done","May 22 2026","NER","","10 days."),
    (4,"QA","NPD Validation & Feedback V4.8 (Non-Basis verification)","Done","Jun 15 2026","NPD","","Non-basis complete."),
    (5,"QA","Complete Build out of Authority again (v4.8)","Planned","Jun 24 2026","NER","Feature 1290619","Blocked on warehouse."),
    (6,"QA","NPD Validation & Feedback V4.8 (Basis)","Planned","Jun 30 2026","NPD","","Requires id 5 complete."),
    (7,"QA","BAHR Authority Build Complete","Planned","Jul 7 2026","NER","",""),
    (8,"QA","Load V5.0 Model","Planned","Jul 7 2026","Dev","",""),
    (9,"QA","NPD Validation V5.0 (Non-Basis)","Planned","Jul 21 2026","NPD","",""),
    (10,"QA","NPD Validation V5.0 (Basis)","Planned","Aug 4 2026","NPD","",""),
    (11,"QA","Load V6.0 Model","Planned","Aug 18 2026","Dev","",""),
    (12,"QA","NPD Validation V6.0 (Non-Basis)","Planned","Sep 1 2026","NPD","",""),
    (13,"QA","NPD Validation V6.0 (Basis)","Planned","Sep 15 2026","NPD","",""),
    (14,"QA","V6.1 Model Load & Validation","Planned","Oct 6 2026","NPD/Dev","",""),
    (15,"QA","V6.2 Model Load & Final QA Validation","Planned","Oct 27 2026","NPD/Dev","",""),
    (16,"QA","QA Sign-Off & UAT Readiness","Planned","Nov 10 2026","QA Lead","",""),
    (17,"QA","Performance Benchmark QA vs Prod Baseline","Planned","Nov 17 2026","Dev","",""),
    (18,"Prod","Prod Infra Provisioning","Planned","Aug 1 2026","Infra","","Requires QA sign-off."),
    (19,"Prod","Data Migration Dry Run #1","Planned","Aug 25 2026","Dev/DBA","",""),
    (20,"Prod","Data Migration Dry Run #2","Planned","Sep 8 2026","Dev/DBA","",""),
    (21,"Prod","Cutover Runbook Draft","Planned","Sep 15 2026","PM","",""),
    (22,"Prod","Load V6.0 to Prod","Planned","Oct 13 2026","Dev","",""),
    (23,"Prod","NPD Prod Spot-Check V6.0","Planned","Oct 20 2026","NPD","",""),
    (24,"Prod","Load V6.1 to Prod","Planned","Nov 3 2026","Dev","",""),
    (25,"Prod","Load V6.2 to Prod","Planned","Nov 17 2026","Dev","",""),
    (26,"Prod","Full Regression Test Prod","Planned","Nov 24 2026","QA/NPD","",""),
    (27,"Prod","Go/No-Go Review","Planned","Dec 1 2026","Leadership","",""),
    (28,"Prod","Production Cutover","Planned","Dec 8 2026","Dev/Infra","","All hands."),
]

for row_idx, row in enumerate(rows_data, 2):
    for col_idx, val in enumerate(row, 1):
        style_data_cell(ws1.cell(row=row_idx, column=col_idx, value=val), row_idx, wrap=(col_idx in (3, 8)))

last_row1 = len(rows_data) + 1
ws1.conditional_formatting.add(f"D2:D{last_row1}", CellIsRule(operator="equal", formula=['"Done"'],        fill=green_fill))
ws1.conditional_formatting.add(f"D2:D{last_row1}", CellIsRule(operator="equal", formula=['"In Progress"'], fill=amber_fill))
ws1.conditional_formatting.add(f"D2:D{last_row1}", CellIsRule(operator="equal", formula=['"Planned"'],     fill=blue_fill))
auto_size_columns(ws1)

# ── Sheet 2: Iterations ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Iterations")
ws2.freeze_panes = "A2"

col_names2 = ["Iter Key","Iter Label","Section","Name","Team","Owner","Health","Status","State","Risks","Next Steps","Impact","Link"]
for col_idx, name in enumerate(col_names2, 1):
    style_header(ws2.cell(row=1, column=col_idx, value=name))

iter_data = [
    ("jun23","Jun 23 2026 (current)","Person Authority Build","Basis Person Authority Build Up","NER","Krishanu Bhattacharya","Green","In Progress","v4.8 basis authority build 60% complete; Warehouse slot confirmed for Jun 24","Team availability during school holidays; Warehouse access window tight","Complete BAHR authority build by Jun 24; Kick off NPD Basis validation Jun 30","Delays here push V5.0 load by 1:1 ratio",""),
    ("jun23","Jun 23 2026 (current)","NPD Validation","Non-Basis V4.8 Validation","NPD","Vandana Roy","Green","Complete","All 200 non-basis records validated; Precision 94.2% / Recall 91.8%","","Archive results to SharePoint; Brief leads on findings","Unlocks V5.0 non-basis track",""),
    ("jun23","Jun 23 2026 (current)","Model Release","V4.8 Drop-In Release","Dev","Dev Team","Amber","In Progress","Model load complete; Post-load checks pending","One flaky integration test — investigating","Resolve integration test; Tag release in ADO","Blocks V5.0 model prep if not resolved",""),
    ("jul07","Jul 7 2026","Person Authority Build","BAHR Authority Build Complete","NER","Krishanu Bhattacharya","Amber","Not Started","Scheduled to start Jun 25","Warehouse availability unconfirmed beyond Jun 24; SME bandwidth — 2 team members on leave","Confirm warehouse slot Jul 1-4; Assign backup SME","Critical path — all V5.0 QA work depends on this",""),
    ("jul07","Jul 7 2026","NPD Validation","Basis V4.8 Validation","NPD","Vandana Roy","Amber","Not Started","Dependent on BAHR build completing Jun 24","If BAHR slips, validation moves to Jul 14 at earliest","Prepare test data set; Confirm NPD reviewer availability","Validation result feeds V5.0 model tuning",""),
    ("jul21","Jul 21 2026","NPD Validation","Non-Basis V5.0 Validation","NPD","Vandana Roy","Green","Not Started","Planned, awaiting V5.0 model load","Model load delay would push this","V5.0 model load Jul 7; Kick off validation immediately after","Feeds V5.0 basis track",""),
]

for row_idx, row in enumerate(iter_data, 2):
    for col_idx, val in enumerate(row, 1):
        style_data_cell(ws2.cell(row=row_idx, column=col_idx, value=val), row_idx, wrap=(col_idx in (9,10,11,12)))

last_row2 = len(iter_data) + 1
ws2.conditional_formatting.add(f"G2:G{last_row2}", CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill))
ws2.conditional_formatting.add(f"G2:G{last_row2}", CellIsRule(operator="equal", formula=['"Amber"'], fill=amber_fill))
ws2.conditional_formatting.add(f"G2:G{last_row2}", CellIsRule(operator="equal", formula=['"Red"'],   fill=red_fill))
auto_size_columns(ws2, max_width=45)

# ── Sheet 3: Files ─────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Files")
ws3.freeze_panes = "A2"
for col_idx, name in enumerate(["File Name","Description","Category","URL","Last Updated","Added By"], 1):
    style_header(ws3.cell(row=1, column=col_idx, value=name))
auto_size_columns(ws3)

output_path = r"C:\Claude\delivery_data.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
